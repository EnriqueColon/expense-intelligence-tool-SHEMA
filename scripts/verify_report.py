"""Verify that the generated Excel report is arithmetically sound.

The workbook ships formulas with no cached results, so the only way to check it
is to have a real spreadsheet engine recalculate the file and then read the
computed values back.  This script builds a report from synthetic records that
exercise the awkward cases (credits, blank categories, unparseable dates,
wildcard characters, multi-month multi-cardholder data), recalculates it with
headless LibreOffice, and asserts the figures tie out.

Run:  python scripts/verify_report.py
"""

import os
import shutil
import subprocess
import sys
import tempfile

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from openpyxl import load_workbook

from app.report import (
    S_CAT, S_DETAIL, S_EXEC, S_TOP, S_TREND, S_VEND,
    EXEC_TOTAL_CHARGES_ROW, EXEC_CREDITS_ROW, EXEC_NET_ROW,
    build_report,
)

ERROR_VALUES = {"#REF!", "#VALUE!", "#DIV/0!", "#NAME?", "#N/A", "#NULL!", "#NUM!", "#ERROR!"}

SOFFICE_CANDIDATES = [
    "/Applications/LibreOffice.app/Contents/MacOS/soffice",
    shutil.which("soffice") or "",
    shutil.which("libreoffice") or "",
]


def find_soffice() -> str:
    for path in SOFFICE_CANDIDATES:
        if path and os.path.exists(path):
            return path
    raise SystemExit("LibreOffice not found — cannot recalculate the workbook.")


def synthetic_records() -> list:
    """Records chosen to stress the parts most likely to break."""
    rows = [
        # id, period,     sale,         post,         description,                 category,               holder,          amount
        (1,  "2026-01", "2026-01-04", "2026-01-05", "AMZN Mktp US*2H43K1",       "Office Supplies",      "Carlos Rivera", 245.19),
        (2,  "2026-01", "2026-01-09", "2026-01-10", "ADOBE  *CREATIVE CLOUD",    "Software",             "Carlos Rivera", 59.99),
        (3,  "2026-01", "2026-01-15", "2026-01-16", "ONLINE PAYMENT THANK YOU",  "Payment",              "Carlos Rivera", -1500.00),
        (4,  "2026-01", "2026-01-21", "2026-01-22", "UBER   *TRIP  HELP.UBER",   "Travel",               "Lauro R Serrano", 38.42),
        (5,  "2026-01", "2026-01-28", "2026-01-29", "STAPLES 00312 AUSTIN TX",   "",                     "Lauro R Serrano", 112.75),
        (6,  "2026-02", "2026-02-02", "2026-02-03", "GODADDY.COM 480-505-8855",  "Dues & Subscriptions", "Carlos Rivera", 219.88),
        (7,  "2026-02", "2026-02-08", "2026-02-09", "ZOOM.US 888-799-9666",      "Telephone/Internet/Web", "Carlos Rivera", 149.90),
        (8,  "2026-02", "2026-02-14", "2026-02-15", "AMAZON WEB SERVICES",       "Computer & Internet",  "Lauro R Serrano", 802.31),
        (9,  "2026-02", "2026-02-19", "2026-02-20", "MERCHANT REFUND",           "Other Expense",        "Lauro R Serrano", -74.20),
        (10, "2026-02", "2026-02-25", "2026-02-26", "FEDEX 782341199",           "Postage & Shipping",   "Carlos Rivera", 46.10),
        # Unparseable date fragment — parser could not read a billing period.
        (11, "2026-03", "03/06",      "03/07",      "FOUR SEASONS HOTEL  NY",    "Travel",               "Carlos Rivera", 1284.00),
        (12, "2026-03", "2026-03-11", "2026-03-12", "MICROSOFT*365 BUSINESS",    "Software",             "Lauro R Serrano", 330.00),
        (13, "2026-03", "2026-03-17", "2026-03-18", "WHAT? A ODD *VENDOR~ NAME", "Meals & Entertainment", "Dana Whitfield", 88.45),
        (14, "2026-03", "2026-03-23", "2026-03-24", "INTUIT *QUICKBOOKS ONLINE", "Professional Fees",    "Dana Whitfield", 90.00),
        (15, "2026-03", "2026-03-30", "2026-03-31", "AMZN Mktp US*9K21LL",       "Office Supplies",      "Dana Whitfield", 421.66),
    ]
    return [
        {
            "id": r[0], "statement_period": r[1], "sale_date": r[2], "post_date": r[3],
            "description": r[4], "category": r[5], "cardholder": r[6], "amount": r[7],
            "batch_id": 100 + (r[0] % 3), "filename": f"statement_{r[1]}.pdf",
            "processed_by": "admin", "notes": "",
        }
        for r in rows
    ]


def recalculate(path: str, outdir: str) -> str:
    """Round-trip the workbook through LibreOffice so formulas are computed."""
    profile = os.path.join(outdir, "lo-profile")
    subprocess.run(
        [
            find_soffice(),
            f"-env:UserInstallation=file://{profile}",
            "--headless", "--norestore",
            "--convert-to", "xlsx", "--outdir", outdir, path,
        ],
        check=True, capture_output=True, timeout=180,
    )
    out = os.path.join(outdir, os.path.basename(path))
    if not os.path.exists(out):
        raise SystemExit("LibreOffice produced no output file.")
    return out


def find_row(ws, label: str, col: int = 1) -> int:
    for row in range(1, ws.max_row + 1):
        if ws.cell(row=row, column=col).value == label:
            return row
    raise AssertionError(f"Row labelled {label!r} not found on {ws.title!r}")


def approx(a, b, tol=0.01) -> bool:
    return a is not None and b is not None and abs(float(a) - float(b)) < tol


def scan_errors(wb) -> list:
    found = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.strip() in ERROR_VALUES:
                    found.append(f"{ws.title}!{cell.coordinate}={cell.value}")
    return found


def edge_scenarios() -> list:
    """Shapes that have historically broken formula layouts."""
    def rec(i, period, cat, holder, amount, desc="VENDOR ONE PURCHASE"):
        return {
            "id": i, "statement_period": period, "sale_date": f"{period}-05",
            "post_date": f"{period}-06", "description": desc, "category": cat,
            "cardholder": holder, "amount": amount, "batch_id": 1,
            "filename": "s.pdf", "processed_by": "admin", "notes": "",
        }

    base = {"start": None, "end": None, "cardholder": None, "period_label": "All",
            "generated_by": "admin", "generated_at": "2026-08-18 12:00"}

    return [
        ("single record", [rec(1, "2026-01", "Software", "Carlos Rivera", 10.0)], base, None),
        # No charges at all — every average and percentage divides by zero.
        ("all credits, no charges",
         [rec(1, "2026-01", "Payment", "Carlos Rivera", -100.0),
          rec(2, "2026-01", "Payment", "Carlos Rivera", -50.0)], base, None),
        # Prior-period block plus a single-cardholder breakdown column.
        ("single month, single cardholder, prior period",
         [rec(i, "2026-02", "Travel", "Dana Whitfield", float(i)) for i in range(1, 6)],
         dict(base, start="2026-02", end="2026-02", cardholder="Dana Whitfield",
              period_label="Feb 2026"), 500.0),
        # Months with no transactions must still produce zero rows, not errors.
        ("gap months in range",
         [rec(1, "2026-01", "Software", "A B", 10.0),
          rec(2, "2026-06", "Travel", "A B", 20.0)], base, None),
    ]


def verify_edge_scenarios(tmp: str) -> list:
    print("\nEdge scenarios (recalculated, checked for error cells)")
    failures = []
    for i, (name, records, params, prior) in enumerate(edge_scenarios()):
        path = os.path.join(tmp, f"edge_{i}.xlsx")
        with open(path, "wb") as fh:
            fh.write(build_report(records, params, prior))
        outdir = os.path.join(tmp, f"edge_{i}_recalc")
        os.makedirs(outdir, exist_ok=True)
        wb = load_workbook(recalculate(path, outdir), data_only=True)
        errors = scan_errors(wb)
        ok = not errors
        print(f"  {'PASS' if ok else 'FAIL'}  {name}"
              f"{'' if ok else ' — ' + '; '.join(errors[:5])}")
        if not ok:
            failures.append(f"edge: {name}")
    return failures


def main() -> int:
    records = synthetic_records()
    expected_charges = sum(r["amount"] for r in records if r["amount"] > 0)
    expected_credits = abs(sum(r["amount"] for r in records if r["amount"] < 0))
    expected_net = sum(r["amount"] for r in records)

    params = {
        "start": "2026-01", "end": "2026-03", "cardholder": None,
        "period_label": "Jan 2026 – Mar 2026",
        "generated_by": "admin", "generated_at": "2026-08-18 12:00",
    }

    tmp = tempfile.mkdtemp(prefix="report-verify-")
    src = os.path.join(tmp, "report.xlsx")
    with open(src, "wb") as fh:
        fh.write(build_report(records, params, prior_total=4820.55))
    print(f"Built workbook: {os.path.getsize(src):,} bytes")

    outdir = os.path.join(tmp, "recalc")
    os.makedirs(outdir, exist_ok=True)
    calc = recalculate(src, outdir)
    print(f"Recalculated via LibreOffice: {calc}")

    wb = load_workbook(calc, data_only=True)
    failures = []

    def check(name: str, ok: bool, detail: str = ""):
        print(f"  {'PASS' if ok else 'FAIL'}  {name}{(' — ' + detail) if detail else ''}")
        if not ok:
            failures.append(name)

    # --- Detail sheet is the source of truth ---------------------------------
    detail = wb[S_DETAIL]
    detail_amounts = [
        detail.cell(row=r, column=9).value
        for r in range(3, 3 + len(records))
    ]
    detail_sum_charges = sum(a for a in detail_amounts if a and a > 0)
    detail_sum_all = sum(a for a in detail_amounts if a)

    print("\nTie-outs")
    check("Detail charges match source data", approx(detail_sum_charges, expected_charges),
          f"{detail_sum_charges:,.2f}")

    # --- Executive Summary ---------------------------------------------------
    exec_ws = wb[S_EXEC]
    exec_total = exec_ws.cell(row=EXEC_TOTAL_CHARGES_ROW, column=2).value
    exec_credits = exec_ws.cell(row=EXEC_CREDITS_ROW, column=2).value
    exec_net = exec_ws.cell(row=EXEC_NET_ROW, column=2).value

    check("Exec Total Charges = sum of positive Detail amounts",
          approx(exec_total, detail_sum_charges), f"{exec_total:,.2f}")
    check("Exec Credits = sum of negative Detail amounts",
          approx(exec_credits, expected_credits), f"{exec_credits:,.2f}")
    check("Exec Net Activity = sum of all Detail amounts",
          approx(exec_net, detail_sum_all), f"{exec_net:,.2f}")

    # --- Monthly breakdown grand total ---------------------------------------
    grand_row = find_row(exec_ws, "Grand Total")
    grand_total = None
    for col in range(exec_ws.max_column, 1, -1):
        value = exec_ws.cell(row=grand_row, column=col).value
        if isinstance(value, (int, float)):
            grand_total = value
            break
    check("Monthly breakdown grand total = Exec Total Charges",
          approx(grand_total, exec_total), f"{grand_total:,.2f}")

    # --- Category Analysis ---------------------------------------------------
    cat_ws = wb[S_CAT]
    cat_total_row = find_row(cat_ws, "Total")
    cat_total = cat_ws.cell(row=cat_total_row, column=3).value
    check("Category Analysis total = Exec Total Charges",
          approx(cat_total, exec_total), f"{cat_total:,.2f}")

    # --- Vendor Analysis -----------------------------------------------------
    ven_ws = wb[S_VEND]
    ven_total_row = find_row(ven_ws, "Total")
    ven_total = ven_ws.cell(row=ven_total_row, column=3).value
    check("Vendor Analysis total = Exec Total Charges",
          approx(ven_total, exec_total), f"{ven_total:,.2f}")

    # Drill-down rows are hidden; confirm their references resolved to numbers.
    drill_rows = [r for r in range(7, ven_total_row)
                  if ven_ws.row_dimensions[r].outlineLevel == 1]
    drill_values = [ven_ws.cell(row=r, column=3).value for r in drill_rows]
    resolved = [v for v in drill_values if isinstance(v, (int, float))]
    check("Drill-down amount cells resolve to numbers",
          len(drill_rows) > 0 and len(resolved) == len(drill_rows),
          f"{len(resolved)}/{len(drill_rows)} rows")
    check("Drill-down amounts sum to Vendor total",
          approx(sum(resolved), exec_total), f"{sum(resolved):,.2f}")

    # --- Monthly Trend -------------------------------------------------------
    trend_ws = wb[S_TREND]
    trend_total_row = find_row(trend_ws, "Total")
    trend_total = trend_ws.cell(row=trend_total_row, column=3).value
    check("Monthly Trend total = Exec Total Charges",
          approx(trend_total, exec_total), f"{trend_total:,.2f}")

    # The category pivot has its own Total row further down.
    pivot_total_row = None
    for row in range(trend_total_row + 1, trend_ws.max_row + 1):
        if trend_ws.cell(row=row, column=1).value == "Total":
            pivot_total_row = row
            break
    if pivot_total_row:
        last_col = trend_ws.max_column
        pivot_total = None
        for col in range(last_col, 1, -1):
            value = trend_ws.cell(row=pivot_total_row, column=col).value
            if isinstance(value, (int, float)):
                pivot_total = value
                break
        check("Monthly-by-category pivot total = Exec Total Charges",
              approx(pivot_total, exec_total), f"{pivot_total:,.2f}")

    # --- Top 25 --------------------------------------------------------------
    top_ws = wb[S_TOP]
    top_amounts = [top_ws.cell(row=r, column=7).value for r in range(7, 7 + len(records))]
    top_resolved = [v for v in top_amounts if isinstance(v, (int, float))]
    check("Top-N reference cells resolve to numbers",
          len(top_resolved) == len(records), f"{len(top_resolved)}/{len(records)}")
    check("Top-N is sorted by absolute amount",
          top_resolved == sorted(top_resolved, key=lambda v: -abs(v)))

    # --- No error cells anywhere ---------------------------------------------
    errors = scan_errors(wb)
    check("Zero formula-error cells", not errors, "; ".join(errors[:5]) or "none")

    # --- Blank-category normalisation ----------------------------------------
    categories = {detail.cell(row=r, column=7).value for r in range(3, 3 + len(records))}
    check("Blank categories normalised to 'Uncategorized'",
          "" not in categories and None not in categories and "Uncategorized" in categories)

    failures.extend(verify_edge_scenarios(tmp))

    print(f"\nWorkbook kept at: {calc}")
    if failures:
        print(f"\n{len(failures)} CHECK(S) FAILED: {', '.join(failures)}")
        return 1
    print("\nAll checks passed.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
