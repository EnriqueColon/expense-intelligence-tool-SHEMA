"""Formula-driven Excel report builder.

The workbook is deliberately *not* a dump of pre-computed numbers.  The last
sheet, "Transaction Detail", holds one row per transaction and is the single
source of truth; every figure on every other sheet is a live formula
(SUMIFS/COUNTIFS/direct cell references) pointing back at it.  Editing or
deleting a row in Detail recalculates the entire workbook.

Because the values are formulas rather than cached results, cells appear blank
until a spreadsheet engine recalculates.  ``fullCalcOnLoad`` makes Excel,
Google Sheets and LibreOffice do that on open; naive preview tools that only
read cached values will show blanks.
"""

from __future__ import annotations

import io
from datetime import datetime, date
from typing import Optional

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import PageSetupProperties

from app.vendors import normalize_vendor, wildcard_safe

BRAND = "SHEMA"

# --- Sheet names -------------------------------------------------------------
S_EXEC   = "Executive Summary"
S_CAT    = "Category Analysis"
S_VEND   = "Vendor Analysis"
S_TREND  = "Monthly Trend"
S_TOP    = "Top 25 Transactions"
S_DETAIL = "Transaction Detail"

# Sheet names containing spaces must be single-quoted inside formula references.
Q_DETAIL = f"'{S_DETAIL}'"
Q_EXEC   = f"'{S_EXEC}'"

# --- Detail sheet geometry ---------------------------------------------------
# Row 1 is a title band, row 2 the header, data begins at row 3.
DETAIL_TITLE_ROW  = 1
DETAIL_HEADER_ROW = 2
DETAIL_FIRST_ROW  = 3

COL_ID, COL_PERIOD, COL_SALE, COL_POST, COL_DESC, COL_VENDOR = 1, 2, 3, 4, 5, 6
COL_CAT, COL_HOLDER, COL_AMOUNT, COL_BATCH, COL_FILE, COL_BY, COL_NOTES = 7, 8, 9, 10, 11, 12, 13

DETAIL_HEADERS = [
    "Txn ID", "Statement Period", "Sale Date", "Post Date", "Description",
    "Vendor", "Category", "Cardholder", "Amount", "Batch ID", "Source File",
    "Processed By", "Notes",
]

# --- Executive Summary geometry ----------------------------------------------
# The blocks above Key Metrics are fixed height, so the KPI cells sit at known
# addresses that other sheets can reference.  ``_write_parameters`` asserts the
# parameters block really does end where this assumes.
EXEC_PARAM_FIRST_ROW = 6
EXEC_PARAM_LAST_ROW  = 12
EXEC_KPI_FIRST_ROW   = 15

EXEC_TOTAL_CHARGES_ROW = EXEC_KPI_FIRST_ROW          # 15
EXEC_CREDITS_ROW       = EXEC_KPI_FIRST_ROW + 1      # 16
EXEC_NET_ROW           = EXEC_KPI_FIRST_ROW + 2      # 17
EXEC_TXN_COUNT_ROW     = EXEC_KPI_FIRST_ROW + 3      # 18
EXEC_CHARGE_COUNT_ROW  = EXEC_KPI_FIRST_ROW + 4      # 19
EXEC_AVG_ROW           = EXEC_KPI_FIRST_ROW + 5      # 20
EXEC_VENDORS_ROW       = EXEC_KPI_FIRST_ROW + 6      # 21
EXEC_CATS_ROW          = EXEC_KPI_FIRST_ROW + 7      # 22
EXEC_HOLDERS_ROW       = EXEC_KPI_FIRST_ROW + 8      # 23
EXEC_TOP_CAT_ROW       = EXEC_KPI_FIRST_ROW + 9      # 24
EXEC_TOP_VENDOR_ROW    = EXEC_KPI_FIRST_ROW + 10     # 25

TOTAL_CHARGES_REF = f"{Q_EXEC}!$B${EXEC_TOTAL_CHARGES_ROW}"

# --- Palette -----------------------------------------------------------------
INK        = "FF0F172A"   # near-black header bands
BAND       = "FF1E293B"   # section title bands
ACCENT     = "FF2563EB"   # single accent colour
ACCENT_SOFT= "FFDBEAFE"
ZEBRA      = "FFF8FAFC"
RULE       = "FFE2E8F0"
MUTED      = "FF64748B"
TOTAL_BG   = "FFF1F5F9"

CHART_ACCENT = "2563EB"   # chart fills take RGB without the alpha prefix

FMT_MONEY = '#,##0.00;[Red](#,##0.00)'
FMT_DATE  = 'yyyy-mm-dd'
FMT_PCT   = '0.0%'
FMT_INT   = '#,##0'

F_TITLE   = Font(name="Calibri", size=18, bold=True, color="FFFFFFFF")
F_SUB     = Font(name="Calibri", size=11, bold=True, color="FFFFFFFF")
F_STAMP   = Font(name="Calibri", size=9,  color="FFFFFFFF")
F_BAND    = Font(name="Calibri", size=10, bold=True, color="FFFFFFFF")
F_HEAD    = Font(name="Calibri", size=10, bold=True, color="FFFFFFFF")
F_LABEL   = Font(name="Calibri", size=10, bold=True, color=INK)
F_BODY    = Font(name="Calibri", size=10, color=INK)
F_MUTED   = Font(name="Calibri", size=9,  italic=True, color=MUTED)
F_TOTAL   = Font(name="Calibri", size=10, bold=True, color=INK)

FILL_INK    = PatternFill("solid", fgColor=INK)
FILL_BAND   = PatternFill("solid", fgColor=BAND)
FILL_ACCENT = PatternFill("solid", fgColor=ACCENT)
FILL_ZEBRA  = PatternFill("solid", fgColor=ZEBRA)
FILL_TOTAL  = PatternFill("solid", fgColor=TOTAL_BG)
FILL_SOFT   = PatternFill("solid", fgColor=ACCENT_SOFT)

_thin = Side(style="thin", color=RULE)
BORDER_ALL = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

MONTH_NAMES = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
               "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]


# ---------------------------------------------------------------------------
# Month helpers ('YYYY-MM' strings)
# ---------------------------------------------------------------------------

def _month_key(period: str) -> tuple:
    y, m = period.split("-")
    return int(y), int(m)


def month_add(period: str, delta: int) -> str:
    y, m = _month_key(period)
    total = y * 12 + (m - 1) + delta
    return f"{total // 12}-{total % 12 + 1:02d}"


def month_span(a: str, b: str) -> int:
    """Inclusive count of months from *a* to *b*."""
    ay, am = _month_key(a)
    by, bm = _month_key(b)
    return (by * 12 + bm) - (ay * 12 + am) + 1


def month_range(a: str, b: str) -> list:
    out, cur = [], a
    for _ in range(max(month_span(a, b), 0)):
        out.append(cur)
        cur = month_add(cur, 1)
    return out


def month_label(period: str) -> str:
    y, m = _month_key(period)
    return f"{MONTH_NAMES[m - 1]} {y}"


# ---------------------------------------------------------------------------
# Small styling helpers
# ---------------------------------------------------------------------------

def _band(ws, row: int, last_col: int, text: str, fill=FILL_BAND, font=F_BAND, height=18):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_col)
    c = ws.cell(row=row, column=1, value=text)
    c.font, c.fill = font, fill
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    for col in range(1, last_col + 1):
        ws.cell(row=row, column=col).fill = fill
    ws.row_dimensions[row].height = height


def _header_row(ws, row: int, headers: list, widths: Optional[list] = None, start_col: int = 1):
    for i, text in enumerate(headers):
        c = ws.cell(row=row, column=start_col + i, value=text)
        c.font, c.fill = F_HEAD, FILL_INK
        c.alignment = Alignment(
            horizontal="left" if i == 0 else "center", vertical="center", wrap_text=True
        )
        c.border = BORDER_ALL
    ws.row_dimensions[row].height = 26
    if widths:
        for i, w in enumerate(widths):
            ws.column_dimensions[get_column_letter(start_col + i)].width = w


def _title_block(ws, last_col: int, subtitle: str, stamp: str):
    """Branded title band used at the top of every sheet."""
    _band(ws, 1, last_col, BRAND, fill=FILL_INK, font=F_TITLE, height=30)
    _band(ws, 2, last_col, subtitle, fill=FILL_INK, font=F_SUB, height=20)
    _band(ws, 3, last_col, stamp, fill=FILL_ACCENT, font=F_STAMP, height=16)


def _print_setup(ws, repeat_row: Optional[int] = None):
    """Landscape, scaled to one page wide, with the header repeated on every page.

    The deliverable is an .xlsx, but these reports get printed and PDF'd, and
    the default setup slices wide tables across pages mid-column.
    """
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.print_options.horizontalCentered = True
    if repeat_row:
        ws.print_title_rows = f"{repeat_row}:{repeat_row}"


def _footnote(ws, row: int, last_col: int, text: str):
    c = ws.cell(row=row, column=1, value=text)
    c.font = F_MUTED
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_col)


def _stripe(ws, row: int, last_col: int, index: int, start_col: int = 1):
    for col in range(start_col, last_col + 1):
        c = ws.cell(row=row, column=col)
        c.border = BORDER_ALL
        if index % 2:
            c.fill = FILL_ZEBRA


# ---------------------------------------------------------------------------
# Record preparation
# ---------------------------------------------------------------------------

def _prepare(records: list) -> list:
    """Normalise raw DB rows and assign each its Detail sheet row number."""
    prepared = []
    for i, r in enumerate(records):
        amount = float(r.get("amount") or 0)
        category = wildcard_safe(str(r.get("category") or "")) or "Uncategorized"
        holder = wildcard_safe(str(r.get("cardholder") or "")) or "Primary"
        prepared.append({
            "row":         DETAIL_FIRST_ROW + i,
            "id":          r.get("id"),
            "period":      r.get("statement_period") or "",
            "sale_date":   r.get("sale_date") or "",
            "post_date":   r.get("post_date") or "",
            "description": r.get("description") or "",
            "vendor":      normalize_vendor(r.get("description") or ""),
            "category":    category,
            "cardholder":  holder,
            "amount":      amount,
            "batch_id":    r.get("batch_id"),
            "filename":    r.get("filename") or "",
            "processed_by": r.get("processed_by") or "",
            "notes":       r.get("notes") or "",
        })
    return prepared


def _as_date(value: str):
    """Return a real date for 'YYYY-MM-DD', else the original string.

    Statements print dates as MM/DD with no year; the parser upgrades them to
    full ISO dates when it can read a billing period, but older rows may still
    hold the raw fragment.
    """
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except (ValueError, TypeError):
        return value


# ---------------------------------------------------------------------------
# Sheet 6 — Transaction Detail (source of truth)
# ---------------------------------------------------------------------------

def _write_detail(ws, rows: list, stamp: str) -> int:
    last_col = len(DETAIL_HEADERS)
    _band(ws, DETAIL_TITLE_ROW, last_col,
          f"{BRAND}  ·  TRANSACTION DETAIL  ·  SOURCE OF TRUTH FOR ALL FORMULAS",
          fill=FILL_INK, font=F_SUB, height=22)
    _header_row(ws, DETAIL_HEADER_ROW, DETAIL_HEADERS,
                widths=[9, 15, 12, 12, 46, 24, 24, 20, 14, 10, 26, 15, 30])

    for i, r in enumerate(rows):
        row = r["row"]
        ws.cell(row=row, column=COL_ID,      value=r["id"])
        ws.cell(row=row, column=COL_PERIOD,  value=r["period"])
        ws.cell(row=row, column=COL_SALE,    value=_as_date(r["sale_date"])).number_format = FMT_DATE
        ws.cell(row=row, column=COL_POST,    value=_as_date(r["post_date"])).number_format = FMT_DATE
        ws.cell(row=row, column=COL_DESC,    value=r["description"])
        ws.cell(row=row, column=COL_VENDOR,  value=r["vendor"])
        ws.cell(row=row, column=COL_CAT,     value=r["category"])
        ws.cell(row=row, column=COL_HOLDER,  value=r["cardholder"])
        ws.cell(row=row, column=COL_AMOUNT,  value=r["amount"]).number_format = FMT_MONEY
        ws.cell(row=row, column=COL_BATCH,   value=r["batch_id"])
        ws.cell(row=row, column=COL_FILE,    value=r["filename"])
        ws.cell(row=row, column=COL_BY,      value=r["processed_by"])
        ws.cell(row=row, column=COL_NOTES,   value=r["notes"])
        for col in range(1, last_col + 1):
            ws.cell(row=row, column=col).font = F_BODY
        _stripe(ws, row, last_col, i)

    last_row = DETAIL_FIRST_ROW + len(rows) - 1
    ws.auto_filter.ref = f"A{DETAIL_HEADER_ROW}:{get_column_letter(last_col)}{last_row}"
    ws.freeze_panes = f"A{DETAIL_FIRST_ROW}"
    ws.sheet_view.showGridLines = False
    _print_setup(ws, repeat_row=DETAIL_HEADER_ROW)
    return last_row


# ---------------------------------------------------------------------------
# Sheet 1 — Executive Summary
# ---------------------------------------------------------------------------

def _write_parameters(ws, params: dict, record_count: int):
    _band(ws, 5, 8, "REPORT PARAMETERS")
    entries = [
        ("Period Start",     params.get("start") or "All"),
        ("Period End",       params.get("end") or "All"),
        ("Cardholder",       params.get("cardholder") or "All Cardholders"),
        ("Categories",       "All"),
        ("Records Included", record_count),
        ("Generated By",     params.get("generated_by") or ""),
        ("Generated At",     params.get("generated_at") or ""),
    ]
    row = EXEC_PARAM_FIRST_ROW
    for label, value in entries:
        lc = ws.cell(row=row, column=1, value=label)
        lc.font = F_LABEL
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        vc = ws.cell(row=row, column=2, value=value)
        vc.font = F_BODY
        vc.alignment = Alignment(horizontal="left", indent=1)
        if label == "Records Included":
            vc.number_format = FMT_INT
        for col in range(1, 5):
            ws.cell(row=row, column=col).border = BORDER_ALL
        row += 1

    # Downstream sheets reference KPI cells by fixed address.
    assert row - 1 == EXEC_PARAM_LAST_ROW, (
        f"Parameters block ended at row {row - 1}, expected {EXEC_PARAM_LAST_ROW}; "
        "EXEC_* row constants need updating."
    )


def _write_exec(ws, rows, ranges, params, months, cardholders, top_category,
                top_vendor, prior_total, stamp):
    ws.sheet_view.showGridLines = False
    for col, width in zip("ABCDEFGH", [30, 22, 18, 16, 16, 16, 16, 16]):
        ws.column_dimensions[col].width = width

    _title_block(ws, 8, "Expense Analysis Report", stamp)
    _write_parameters(ws, params, len(rows))

    R_AMT, R_CAT, R_VEN, R_CH, R_PER = (
        ranges["amount"], ranges["category"], ranges["vendor"],
        ranges["cardholder"], ranges["period"],
    )

    _band(ws, 14, 8, "KEY METRICS")

    def kpi(row, label, formula, fmt=FMT_MONEY):
        lc = ws.cell(row=row, column=1, value=label)
        lc.font = F_LABEL
        lc.border = BORDER_ALL
        vc = ws.cell(row=row, column=2, value=formula)
        vc.font = F_BODY
        vc.number_format = fmt
        vc.border = BORDER_ALL
        vc.alignment = Alignment(horizontal="right")
        return vc

    kpi(EXEC_TOTAL_CHARGES_ROW, "Total Charges",       f'=SUMIF({R_AMT},">0")')
    kpi(EXEC_CREDITS_ROW,       "Credits & Payments",  f'=ABS(SUMIF({R_AMT},"<0"))')
    kpi(EXEC_NET_ROW,           "Net Activity",        f'=SUM({R_AMT})')
    kpi(EXEC_TXN_COUNT_ROW,     "Transactions",        f'=COUNT({R_AMT})', FMT_INT)
    kpi(EXEC_CHARGE_COUNT_ROW,  "Charge Transactions", f'=COUNTIF({R_AMT},">0")', FMT_INT)
    kpi(EXEC_AVG_ROW,           "Average Charge",
        f'=IFERROR($B${EXEC_TOTAL_CHARGES_ROW}/$B${EXEC_CHARGE_COUNT_ROW},0)')
    kpi(EXEC_VENDORS_ROW,  "Distinct Vendors",
        f'=SUMPRODUCT(({R_VEN}<>"")/COUNTIF({R_VEN},{R_VEN}&""))', FMT_INT)
    kpi(EXEC_CATS_ROW,     "Distinct Categories",
        f'=SUMPRODUCT(({R_CAT}<>"")/COUNTIF({R_CAT},{R_CAT}&""))', FMT_INT)
    kpi(EXEC_HOLDERS_ROW,  "Distinct Cardholders",
        f'=SUMPRODUCT(({R_CH}<>"")/COUNTIF({R_CH},{R_CH}&""))', FMT_INT)

    # Top category / vendor: the label is a static value, the spend beside it is
    # a SUMIFS driven by that label cell.
    for row, label, name, rng in (
        (EXEC_TOP_CAT_ROW,    "Top Category", top_category, R_CAT),
        (EXEC_TOP_VENDOR_ROW, "Top Vendor",   top_vendor,   R_VEN),
    ):
        lc = ws.cell(row=row, column=1, value=label)
        lc.font, lc.border = F_LABEL, BORDER_ALL
        nc = ws.cell(row=row, column=2, value=name)
        nc.font, nc.border = F_BODY, BORDER_ALL
        nc.alignment = Alignment(horizontal="right")
        sc = ws.cell(row=row, column=3,
                     value=f'=SUMIFS({R_AMT},{rng},$B${row},{R_AMT},">0")')
        sc.font, sc.border = F_BODY, BORDER_ALL
        sc.number_format = FMT_MONEY

    # --- Monthly breakdown by cardholder ------------------------------------
    row = EXEC_TOP_VENDOR_ROW + 2
    last_col = 2 + len(cardholders)
    _band(ws, row, max(last_col, 8), "MONTHLY BREAKDOWN BY CARDHOLDER")
    header_row = row + 1
    _header_row(ws, header_row, ["Month"] + cardholders + ["Total"])
    for i in range(len(cardholders)):
        ws.column_dimensions[get_column_letter(2 + i)].width = 18
    ws.column_dimensions[get_column_letter(last_col)].width = 18

    first_month_row = header_row + 1
    for i, period in enumerate(months):
        r = first_month_row + i
        mc = ws.cell(row=r, column=1, value=period)
        mc.font = F_BODY
        for j in range(len(cardholders)):
            col = 2 + j
            letter = get_column_letter(col)
            cell = ws.cell(
                row=r, column=col,
                value=(f'=SUMIFS({R_AMT},{R_PER},$A{r},{R_CH},{letter}${header_row},'
                       f'{R_AMT},">0")'),
            )
            cell.number_format = FMT_MONEY
            cell.font = F_BODY
        # Total is an independent SUMIFS on the month alone, so the column ties
        # to Total Charges even if the cardholder columns were ever truncated.
        tc = ws.cell(row=r, column=last_col,
                     value=f'=SUMIFS({R_AMT},{R_PER},$A{r},{R_AMT},">0")')
        tc.number_format = FMT_MONEY
        tc.font = F_TOTAL
        _stripe(ws, r, last_col, i)

    last_month_row = first_month_row + len(months) - 1
    grand_row = last_month_row + 1
    gc = ws.cell(row=grand_row, column=1, value="Grand Total")
    gc.font = F_TOTAL
    for col in range(2, last_col + 1):
        letter = get_column_letter(col)
        c = ws.cell(row=grand_row, column=col,
                    value=f'=SUM({letter}{first_month_row}:{letter}{last_month_row})')
        c.number_format = FMT_MONEY
        c.font = F_TOTAL
    for col in range(1, last_col + 1):
        ws.cell(row=grand_row, column=col).fill = FILL_TOTAL
        ws.cell(row=grand_row, column=col).border = BORDER_ALL

    note_row = grand_row + 1
    _footnote(ws, note_row, max(last_col, 8),
              "Charges only (amount > 0); credits and payments are excluded from this "
              "table and shown separately above.")

    # --- vs. prior period ----------------------------------------------------
    if prior_total is not None:
        row = note_row + 2
        _band(ws, row, 8, "VS. PRIOR PERIOD")
        pr = row + 1
        entries = [
            ("Prior Period Total (static)", prior_total, FMT_MONEY),
            ("Current Period Total", f"={TOTAL_CHARGES_REF}", FMT_MONEY),
        ]
        for i, (label, value, fmt) in enumerate(entries):
            lc = ws.cell(row=pr + i, column=1, value=label)
            lc.font, lc.border = F_LABEL, BORDER_ALL
            vc = ws.cell(row=pr + i, column=2, value=value)
            vc.font, vc.border = F_BODY, BORDER_ALL
            vc.number_format = fmt
            vc.alignment = Alignment(horizontal="right")

        prior_cell, cur_cell = f"$B${pr}", f"$B${pr + 1}"
        deltas = [
            ("Variance ($)", f"={cur_cell}-{prior_cell}", FMT_MONEY),
            ("Variance (%)", f'=IFERROR(({cur_cell}-{prior_cell})/{prior_cell},"n/a")', FMT_PCT),
        ]
        for i, (label, value, fmt) in enumerate(deltas):
            r = pr + 2 + i
            lc = ws.cell(row=r, column=1, value=label)
            lc.font, lc.border = F_LABEL, BORDER_ALL
            vc = ws.cell(row=r, column=2, value=value)
            vc.font, vc.border = F_TOTAL, BORDER_ALL
            vc.number_format = fmt
            vc.alignment = Alignment(horizontal="right")
            vc.fill = FILL_SOFT

        _footnote(ws, pr + 4, 8,
                  "Prior-period total is a static figure: those transactions fall outside "
                  "this report's Detail sheet, so it cannot be a formula.")

    _print_setup(ws)
    return {"grand_total_cell": f"{get_column_letter(last_col)}{grand_row}"}


# ---------------------------------------------------------------------------
# Sheet 2 — Category Analysis
# ---------------------------------------------------------------------------

def _write_categories(ws, ranges, categories, stamp):
    ws.sheet_view.showGridLines = False
    _title_block(ws, 5, "Category Analysis", stamp)
    _band(ws, 5, 5, "SPEND BY CATEGORY")
    header = 6
    _header_row(ws, header, ["Category", "Transactions", "Total Charges", "Average", "% of Total"],
                widths=[34, 14, 18, 16, 14])

    R_AMT, R_CAT = ranges["amount"], ranges["category"]
    first = header + 1
    total_row = first + len(categories)

    for i, name in enumerate(categories):
        r = first + i
        ws.cell(row=r, column=1, value=name).font = F_BODY
        ws.cell(row=r, column=2,
                value=f'=COUNTIFS({R_CAT},$A{r},{R_AMT},">0")').number_format = FMT_INT
        ws.cell(row=r, column=3,
                value=f'=SUMIFS({R_AMT},{R_CAT},$A{r},{R_AMT},">0")').number_format = FMT_MONEY
        ws.cell(row=r, column=4, value=f"=IFERROR(C{r}/B{r},0)").number_format = FMT_MONEY
        ws.cell(row=r, column=5,
                value=f"=IFERROR(C{r}/$C${total_row},0)").number_format = FMT_PCT
        for col in range(2, 6):
            ws.cell(row=r, column=col).font = F_BODY
        _stripe(ws, r, 5, i)

    last = total_row - 1
    ws.cell(row=total_row, column=1, value="Total").font = F_TOTAL
    ws.cell(row=total_row, column=2, value=f"=SUM(B{first}:B{last})").number_format = FMT_INT
    ws.cell(row=total_row, column=3, value=f"=SUM(C{first}:C{last})").number_format = FMT_MONEY
    ws.cell(row=total_row, column=4,
            value=f"=IFERROR(C{total_row}/B{total_row},0)").number_format = FMT_MONEY
    ws.cell(row=total_row, column=5,
            value=f"=IFERROR(C{total_row}/$C${total_row},0)").number_format = FMT_PCT
    for col in range(1, 6):
        c = ws.cell(row=total_row, column=col)
        c.font, c.fill, c.border = F_TOTAL, FILL_TOTAL, BORDER_ALL

    # Horizontal bars carry their own labels, so axis titles would only add noise.
    chart = BarChart()
    chart.type = "bar"
    chart.style = 10
    chart.title = "Total Charges by Category"
    data = Reference(ws, min_col=3, min_row=header, max_row=last)
    cats = Reference(ws, min_col=1, min_row=first, max_row=last)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.legend = None
    chart.series[0].graphicalProperties = GraphicalProperties(solidFill=CHART_ACCENT)
    chart.height, chart.width = max(8, 0.45 * len(categories) + 5), 20
    ws.add_chart(chart, f"G{header}")

    _footnote(ws, total_row + 1, 5,
              "Charges only (amount > 0). Categories showing zero appear on credit or "
              "payment rows, which are excluded from this table.")
    ws.freeze_panes = f"A{first}"
    _print_setup(ws, repeat_row=header)
    return total_row


# ---------------------------------------------------------------------------
# Sheet 3 — Vendor Analysis (with per-vendor drill-down)
# ---------------------------------------------------------------------------

def _write_vendors(ws, ranges, vendor_names, vendor_rows, stamp):
    ws.sheet_view.showGridLines = False
    # Summary row sits ABOVE its detail group.
    ws.sheet_properties.outlinePr.summaryBelow = False

    _title_block(ws, 5, "Vendor Analysis", stamp)
    _band(ws, 5, 5, "SPEND BY VENDOR  ·  EXPAND A VENDOR TO SEE ITS TRANSACTIONS")
    header = 6
    _header_row(ws, header, ["Vendor", "Transactions", "Total Charges", "Average", "% of Total"],
                widths=[38, 14, 18, 16, 14])

    R_AMT, R_VEN = ranges["amount"], ranges["vendor"]
    r = header + 1
    summary_rows = []

    for i, name in enumerate(vendor_names):
        ws.cell(row=r, column=1, value=name).font = F_LABEL
        ws.cell(row=r, column=2,
                value=f'=COUNTIFS({R_VEN},$A{r},{R_AMT},">0")').number_format = FMT_INT
        ws.cell(row=r, column=3,
                value=f'=SUMIFS({R_AMT},{R_VEN},$A{r},{R_AMT},">0")').number_format = FMT_MONEY
        ws.cell(row=r, column=4, value=f"=IFERROR(C{r}/B{r},0)").number_format = FMT_MONEY
        ws.cell(row=r, column=5,
                value=f"=IFERROR(C{r}/{TOTAL_CHARGES_REF},0)").number_format = FMT_PCT
        for col in range(1, 6):
            c = ws.cell(row=r, column=col)
            c.border = BORDER_ALL
            if col > 1:
                c.font = F_BODY
            if i % 2:
                c.fill = FILL_ZEBRA
        summary_rows.append(r)
        r += 1

        # Drill-down rows reference the Detail sheet cell-for-cell.
        for txn in vendor_rows[name]:
            d = txn["row"]
            ws.cell(row=r, column=1, value=f"={Q_DETAIL}!$C${d}").number_format = FMT_DATE
            ws.cell(row=r, column=2, value=f"={Q_DETAIL}!$E${d}")
            ws.cell(row=r, column=3, value=f"={Q_DETAIL}!$I${d}").number_format = FMT_MONEY
            ws.cell(row=r, column=4, value=f"={Q_DETAIL}!$H${d}")
            ws.cell(row=r, column=5, value=f"={Q_DETAIL}!$G${d}")
            for col in range(1, 6):
                c = ws.cell(row=r, column=col)
                c.font = F_BODY
                c.border = BORDER_ALL
                c.alignment = Alignment(indent=2) if col == 1 else Alignment()
            ws.row_dimensions[r].outlineLevel = 1
            ws.row_dimensions[r].hidden = True
            r += 1

    total_row = r
    ws.cell(row=total_row, column=1, value="Total").font = F_TOTAL
    ws.cell(row=total_row, column=2, value=f'=COUNTIF({R_AMT},">0")').number_format = FMT_INT
    ws.cell(row=total_row, column=3, value=f'=SUMIF({R_AMT},">0")').number_format = FMT_MONEY
    ws.cell(row=total_row, column=4,
            value=f"=IFERROR(C{total_row}/B{total_row},0)").number_format = FMT_MONEY
    ws.cell(row=total_row, column=5,
            value=f"=IFERROR(C{total_row}/{TOTAL_CHARGES_REF},0)").number_format = FMT_PCT
    for col in range(1, 6):
        c = ws.cell(row=total_row, column=col)
        c.font, c.fill, c.border = F_TOTAL, FILL_TOTAL, BORDER_ALL

    _footnote(ws, total_row + 1, 5,
              "Charges only (amount > 0). Drill-down rows are collapsed by default and "
              "reference the Transaction Detail sheet directly. Native outline grouping "
              "works in Excel and Google Sheets; Apple Numbers ignores it and shows all rows.")

    # Summary rows are no longer contiguous once drill-downs are interleaved, so
    # a chart cannot reference them. Build a hidden helper block instead.
    helper_col, helper_val = 27, 28   # AA / AB
    top10 = vendor_names[:10]
    ws.cell(row=header, column=helper_col, value="Vendor")
    ws.cell(row=header, column=helper_val, value="Total Charges")
    for i, name in enumerate(top10):
        hr = header + 1 + i
        ws.cell(row=hr, column=helper_col, value=name)
        ws.cell(row=hr, column=helper_val,
                value=f'=SUMIFS({R_AMT},{R_VEN},${get_column_letter(helper_col)}{hr},'
                      f'{R_AMT},">0")').number_format = FMT_MONEY
    ws.column_dimensions[get_column_letter(helper_col)].hidden = True
    ws.column_dimensions[get_column_letter(helper_val)].hidden = True

    if top10:
        chart = BarChart()
        chart.type = "bar"
        chart.style = 10
        chart.title = "Top 10 Vendors by Total Charges"
        data = Reference(ws, min_col=helper_val, min_row=header, max_row=header + len(top10))
        cats = Reference(ws, min_col=helper_col, min_row=header + 1, max_row=header + len(top10))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.legend = None
        chart.series[0].graphicalProperties = GraphicalProperties(solidFill=CHART_ACCENT)
        # The helper range is hidden, and charts skip hidden cells by default —
        # without this the chart renders empty.
        chart.visible_cells_only = False
        chart.height, chart.width = 10, 20
        ws.add_chart(chart, f"G{header}")

    ws.freeze_panes = f"A{header + 1}"
    _print_setup(ws, repeat_row=header)


# ---------------------------------------------------------------------------
# Sheet 4 — Monthly Trend
# ---------------------------------------------------------------------------

def _write_trend(ws, ranges, months, top_categories, stamp):
    ws.sheet_view.showGridLines = False
    _title_block(ws, 8, "Monthly Trend", stamp)

    R_AMT, R_PER, R_CAT = ranges["amount"], ranges["period"], ranges["category"]

    _band(ws, 5, 3, "MONTHLY TOTALS")
    header = 6
    _header_row(ws, header, ["Month", "Transactions", "Total Charges"], widths=[16, 14, 18])
    first = header + 1
    for i, period in enumerate(months):
        r = first + i
        ws.cell(row=r, column=1, value=period).font = F_BODY
        ws.cell(row=r, column=2,
                value=f'=COUNTIFS({R_PER},$A{r},{R_AMT},">0")').number_format = FMT_INT
        ws.cell(row=r, column=3,
                value=f'=SUMIFS({R_AMT},{R_PER},$A{r},{R_AMT},">0")').number_format = FMT_MONEY
        for col in range(2, 4):
            ws.cell(row=r, column=col).font = F_BODY
        _stripe(ws, r, 3, i)
    last = first + len(months) - 1
    total_row = last + 1
    ws.cell(row=total_row, column=1, value="Total").font = F_TOTAL
    ws.cell(row=total_row, column=2, value=f"=SUM(B{first}:B{last})").number_format = FMT_INT
    ws.cell(row=total_row, column=3, value=f"=SUM(C{first}:C{last})").number_format = FMT_MONEY
    for col in range(1, 4):
        c = ws.cell(row=total_row, column=col)
        c.font, c.fill, c.border = F_TOTAL, FILL_TOTAL, BORDER_ALL

    line = LineChart()
    line.title = "Total Charges by Month"
    line.style = 12
    line.y_axis.title = "Total Charges"
    line.x_axis.title = "Statement Month"
    data = Reference(ws, min_col=3, min_row=header, max_row=last)
    cats = Reference(ws, min_col=1, min_row=first, max_row=last)
    line.add_data(data, titles_from_data=True)
    line.set_categories(cats)
    line.series[0].graphicalProperties = GraphicalProperties(
        ln=LineProperties(solidFill=CHART_ACCENT, w=28000)
    )
    line.series[0].smooth = False
    line.height, line.width = 9, 22
    ws.add_chart(line, f"E{header}")

    # --- Monthly by category pivot ------------------------------------------
    pivot_band = total_row + 22
    n_cats = len(top_categories)
    last_col = 2 + n_cats           # Month + categories + Other, then Total
    total_col = last_col + 1
    _band(ws, pivot_band, total_col, "MONTHLY BY CATEGORY (TOP 6)")
    pheader = pivot_band + 1
    _header_row(ws, pheader, ["Month"] + top_categories + ["All Other", "Total"])
    for i in range(1, total_col):
        ws.column_dimensions[get_column_letter(i + 1)].width = 18

    pfirst = pheader + 1
    for i, period in enumerate(months):
        r = pfirst + i
        ws.cell(row=r, column=1, value=period).font = F_BODY
        for j in range(n_cats):
            col = 2 + j
            letter = get_column_letter(col)
            c = ws.cell(row=r, column=col,
                        value=(f'=SUMIFS({R_AMT},{R_PER},$A{r},{R_CAT},{letter}${pheader},'
                               f'{R_AMT},">0")'))
            c.number_format, c.font = FMT_MONEY, F_BODY
        other_letter_first = get_column_letter(2)
        other_letter_last = get_column_letter(1 + n_cats)
        oc = ws.cell(row=r, column=last_col,
                     value=(f'=SUMIFS({R_AMT},{R_PER},$A{r},{R_AMT},">0")'
                            f'-SUM({other_letter_first}{r}:{other_letter_last}{r})'))
        oc.number_format, oc.font = FMT_MONEY, F_BODY
        tc = ws.cell(row=r, column=total_col,
                     value=f'=SUMIFS({R_AMT},{R_PER},$A{r},{R_AMT},">0")')
        tc.number_format, tc.font = FMT_MONEY, F_TOTAL
        _stripe(ws, r, total_col, i)

    plast = pfirst + len(months) - 1
    ptotal = plast + 1
    ws.cell(row=ptotal, column=1, value="Total").font = F_TOTAL
    for col in range(2, total_col + 1):
        letter = get_column_letter(col)
        c = ws.cell(row=ptotal, column=col, value=f"=SUM({letter}{pfirst}:{letter}{plast})")
        c.number_format, c.font = FMT_MONEY, F_TOTAL
    for col in range(1, total_col + 1):
        c = ws.cell(row=ptotal, column=col)
        c.fill, c.border = FILL_TOTAL, BORDER_ALL

    stacked = BarChart()
    stacked.type = "col"
    stacked.grouping = "stacked"
    stacked.overlap = 100
    stacked.title = "Monthly Spend by Category (Top 6 + Other)"
    stacked.y_axis.title = "Total Charges"
    stacked.x_axis.title = "Statement Month"
    sdata = Reference(ws, min_col=2, max_col=last_col, min_row=pheader, max_row=plast)
    scats = Reference(ws, min_col=1, min_row=pfirst, max_row=plast)
    stacked.add_data(sdata, titles_from_data=True)
    stacked.set_categories(scats)
    stacked.height, stacked.width = 10, 24
    ws.add_chart(stacked, f"{get_column_letter(total_col + 2)}{pheader}")

    _footnote(ws, ptotal + 1, total_col,
              "Charges only (amount > 0), grouped by statement billing period rather than "
              "upload date. 'All Other' is the month total less the six categories shown.")
    ws.freeze_panes = f"A{first}"
    _print_setup(ws, repeat_row=header)
    return total_row


# ---------------------------------------------------------------------------
# Sheet 5 — Top 25 Transactions
# ---------------------------------------------------------------------------

def _write_top(ws, top_rows, stamp):
    ws.sheet_view.showGridLines = False
    _title_block(ws, 8, "Top 25 Transactions", stamp)
    _band(ws, 5, 8, "LARGEST TRANSACTIONS BY ABSOLUTE AMOUNT")
    header = 6
    _header_row(ws, header,
                ["Rank", "Date", "Description", "Vendor", "Category", "Cardholder",
                 "Amount", "% of Total Charges"],
                widths=[7, 12, 46, 24, 24, 20, 16, 16])

    first = header + 1
    for i, txn in enumerate(top_rows):
        r = first + i
        d = txn["row"]
        ws.cell(row=r, column=1, value=i + 1).font = F_BODY
        ws.cell(row=r, column=2, value=f"={Q_DETAIL}!$C${d}").number_format = FMT_DATE
        ws.cell(row=r, column=3, value=f"={Q_DETAIL}!$E${d}")
        ws.cell(row=r, column=4, value=f"={Q_DETAIL}!$F${d}")
        ws.cell(row=r, column=5, value=f"={Q_DETAIL}!$G${d}")
        ws.cell(row=r, column=6, value=f"={Q_DETAIL}!$H${d}")
        ws.cell(row=r, column=7, value=f"={Q_DETAIL}!$I${d}").number_format = FMT_MONEY
        ws.cell(row=r, column=8,
                value=f"=IFERROR(G{r}/{TOTAL_CHARGES_REF},0)").number_format = FMT_PCT
        for col in range(1, 9):
            ws.cell(row=r, column=col).font = F_BODY
        _stripe(ws, r, 8, i)

    _footnote(ws, first + len(top_rows) + 1, 8,
              "Ranked by absolute amount, so large credits and payments appear alongside "
              "large charges. Every cell references the Transaction Detail sheet.")
    ws.freeze_panes = f"A{first}"
    _print_setup(ws, repeat_row=header)


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def build_report(records: list, params: dict, prior_total: Optional[float] = None) -> bytes:
    """Build the workbook and return it as .xlsx bytes.

    *records* are raw rows from the shared dashboard query; *params* carries the
    filters plus display metadata; *prior_total* is the static prior-period
    charge total, or None when no start month was selected.
    """
    rows = _prepare(records)
    if not rows:
        raise ValueError("No records to report on.")

    generated_at = params.get("generated_at") or datetime.now().strftime("%Y-%m-%d %H:%M")
    period_label = params.get("period_label") or "All periods"
    stamp = f"Prepared {generated_at}  ·  Period: {period_label}"

    wb = Workbook()
    wb.remove(wb.active)
    ws_exec   = wb.create_sheet(S_EXEC)
    ws_cat    = wb.create_sheet(S_CAT)
    ws_vend   = wb.create_sheet(S_VEND)
    ws_trend  = wb.create_sheet(S_TREND)
    ws_top    = wb.create_sheet(S_TOP)
    ws_detail = wb.create_sheet(S_DETAIL)

    last_row = _write_detail(ws_detail, rows, stamp)

    def rng(col: int) -> str:
        letter = get_column_letter(col)
        return f"{Q_DETAIL}!${letter}${DETAIL_FIRST_ROW}:${letter}${last_row}"

    # Every SUMIF/SUMIFS range spans exactly the same rows.
    ranges = {
        "amount":     rng(COL_AMOUNT),
        "category":   rng(COL_CAT),
        "vendor":     rng(COL_VENDOR),
        "cardholder": rng(COL_HOLDER),
        "period":     rng(COL_PERIOD),
    }

    charges = [r for r in rows if r["amount"] > 0]

    def totals_by(key: str) -> dict:
        out: dict = {}
        for r in charges:
            out[r[key]] = out.get(r[key], 0.0) + r["amount"]
        return out

    cat_totals = totals_by("category")
    ven_totals = totals_by("vendor")
    categories = sorted(cat_totals, key=lambda k: -cat_totals[k])
    vendor_names = sorted(ven_totals, key=lambda k: -ven_totals[k])

    # Categories present only on credit rows still deserve a line.
    for r in rows:
        if r["category"] not in categories:
            categories.append(r["category"])

    cardholders = sorted({r["cardholder"] for r in rows})

    periods = sorted({r["period"] for r in rows if r["period"]})
    if periods:
        lo, hi = periods[0], periods[-1]
        if params.get("start") and params["start"] < lo:
            lo = params["start"]
        if params.get("end") and params["end"] > hi:
            hi = params["end"]
        months = month_range(lo, hi)
    else:
        months = []

    vendor_rows = {name: [] for name in vendor_names}
    for r in charges:
        vendor_rows[r["vendor"]].append(r)

    top_category = categories[0] if categories else ""
    top_vendor = vendor_names[0] if vendor_names else ""
    top_rows = sorted(rows, key=lambda r: -abs(r["amount"]))[:25]

    _write_exec(ws_exec, rows, ranges, params, months, cardholders,
                top_category, top_vendor, prior_total, stamp)
    _write_categories(ws_cat, ranges, categories, stamp)
    _write_vendors(ws_vend, ranges, vendor_names, vendor_rows, stamp)
    _write_trend(ws_trend, ranges, months, categories[:6], stamp)
    _write_top(ws_top, top_rows, stamp)

    # Formulas carry no cached results, so tell the spreadsheet app to compute
    # everything the moment the file opens.
    wb.calculation.fullCalcOnLoad = True

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
